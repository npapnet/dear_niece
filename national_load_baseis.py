"""
Loader for raw baseis xlsx files from data/baseis-raw/.

Each file has a 2-row merged-cell header (rows 2-3) preceded by a title row (row 1).
The column count differs by year (12 cols for 2023/2024, 14 for 2025+).
This module handles all variants automatically.
"""

import pathlib
import re
import openpyxl
import pandas as pd

COLUMN_MAP = {
    'ΚΩΔΙΚΟΣ ΣΧΟΛΗΣ':             'school_code',
    'ΙΔΡΥΜΑ':                      'institution',
    'ΟΝΟΜΑ ΣΧΟΛΗΣ':                'department',
    'ΕΙΔΟΣ ΘΕΣΗΣ':                 'position_type',
    'ΕΠΙΣΤΗΜΟΝΙΚΑ ΠΕΔΙΑ':          '_field_raw',
    'ΑΡΧΙΚΕΣ ΘΕΣΕΙΣ':             'initial_slots',
    'ΘΕΣΕΙΣ (Κατόπιν Μεταφοράς)': 'slots',
    'ΕΠΙΤ/ΤΕΣ':                    'admitted',
    'ΚΕΝΑ':                        'vacancies',
    'ΒΑΘΜΟΣ ΠΡΩΤΟΥ_ΜΟΡΙΑ':        'top_score',
    'ΒΑΘΜΟΣ ΤΕΛΕΥΤΑΙΟΥ_ΜΟΡΙΑ':    'entry',
}

# Columns that carry no usable information and are dropped on load.
_DROP_COLUMNS = {
    'ΙΣΟΒΑΘ.',
    'ΒΑΘΜΟΣ ΠΡΩΤΟΥ_ΚΡΙΤΗΡΙΑ ΙΣΟΒΑΘΜΙΑΣ',
    'ΒΑΘΜΟΣ ΤΕΛΕΥΤΑΙΟΥ_ΚΡΙΤΗΡΙΑ ΙΣΟΒΑΘΜΙΑΣ',
}

_FIELDS = [1, 2, 3, 4]


def _expand_field(df: pd.DataFrame) -> pd.DataFrame:
    """
    Replace the raw `_field_raw` string (e.g. '2/3') with four boolean
    columns `field_1` … `field_4`, one per scientific field.
    Values like '1/2/3/4' set all four; NaN rows get False everywhere.
    """
    raw = df['_field_raw'].fillna('').astype(str)
    for f in _FIELDS:
        df[f'field_{f}'] = raw.str.split('/').apply(lambda parts: str(f) in parts)
    return df.drop(columns=['_field_raw'])


def _build_columns(ws) -> list[str]:
    """
    Construct column names from the 2-row merged header (rows 2 and 3).

    Row 2 holds the main header; some cells span 2 rows (standalone) while
    others span 2 columns and are split into sub-headers in row 3.
    Forward-filling row 2 across None (merged-column) cells lets us pair
    each row-3 sub-header with its parent label.
    """
    row2 = [c.value for c in ws[2]]
    row3 = [c.value for c in ws[3]]

    last = None
    row2_filled = []
    for v in row2:
        if v is not None:
            last = v
        row2_filled.append(last)

    cols = []
    for r2, r3 in zip(row2_filled, row3):
        cols.append(f'{r2}_{r3}' if r3 else (r2 or f'_col{len(cols)}'))
    return cols


def _extract_year(ws) -> int:
    """Read the year from the title cell (row 1, always ends with the 4-digit year)."""
    title = ws.cell(1, 1).value or ''
    match = re.search(r'(\d{4})', title)
    if not match:
        raise ValueError(f'Cannot extract year from title: {title!r}')
    return int(match.group(1))


def load_baseis_raw(filepath: str | pathlib.Path) -> pd.DataFrame:
    """
    Load one raw baseis xlsx file into a tidy DataFrame.

    Returns a DataFrame with English column names (see COLUMN_MAP) plus a
    `year` column extracted from the file title.  Columns not present in a
    given year (e.g. `vacancies` before 2025) are silently absent.
    Fully-empty rows are dropped.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    year = _extract_year(ws)
    cols = _build_columns(ws)

    data = [
        tuple(c.value for c in row)
        for row in ws.iter_rows(min_row=4)
    ]
    df = pd.DataFrame(data, columns=cols)
    df = df.dropna(how='all')

    df.drop(columns=[c for c in _DROP_COLUMNS if c in df.columns], inplace=True)
    df.rename(columns=COLUMN_MAP, inplace=True)
    df = _expand_field(df)
    df['year'] = year
    return df


def build_master(raw_dir: str | pathlib.Path) -> pd.DataFrame:
    """
    Load all gel-*.xlsx files in `raw_dir` and concatenate into one DataFrame.

    The result is in long format: one row per (year, school, position_type).
    Columns that exist only in some years (e.g. `vacancies`) are NaN for
    years that predate them.
    """
    raw_dir = pathlib.Path(raw_dir)
    files = sorted(raw_dir.glob('gel-*.xlsx'))
    frames = []
    for i, f in enumerate(files, 1):
        year = f.stem.split('-')[1]
        print(f'  {year}  ({i}/{len(files)})', end='\n', flush=False)
        frames.append(load_baseis_raw(f))
    print()
    master = pd.concat(frames, ignore_index=True)
    # put year first for readability
    cols = ['year'] + [c for c in master.columns if c != 'year']
    return master[cols]


if __name__ == '__main__':
    import sys
    raw_dir = pathlib.Path(__file__).parent / 'data' / 'baseis-raw'
    master = build_master(raw_dir)
    print(f'Loaded {len(master):,} rows across years: {sorted(master["year"].unique())}')
    out = pathlib.Path(__file__).parent / 'data' / 'baseis-master.csv'
    master.to_csv(out, index=False, encoding='utf-8-sig')
    print(f'Saved → {out}')
